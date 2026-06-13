from rest_framework import generics, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser
from django.db.models import Count, Q
from django.http import HttpResponse
from .models import Student, Class, Section, ClassMaster, ClassSectionMaster
from .serializers import (
    StudentSerializer, StudentListSerializer, ClassSerializer,
    SectionSerializer, ClassStrengthSerializer,
    ClassMasterSerializer, ClassSectionMasterSerializer
)
from utils.permissions import IsSchoolAdmin, IsSchoolStaff


class ClassListCreateView(generics.ListCreateAPIView):
    serializer_class = ClassSerializer
    permission_classes = [IsSchoolStaff]

    def get_queryset(self):
        return Class.objects.prefetch_related('sections').all()


class ClassDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = ClassSerializer
    permission_classes = [IsSchoolAdmin]
    queryset = Class.objects.all()


# ClassMaster views (from fee-backend tables)
class ClassMasterListView(generics.ListAPIView):
    """List classes from class_master table (fee-backend)"""
    serializer_class = ClassMasterSerializer
    permission_classes = [IsSchoolStaff]

    def get_queryset(self):
        return ClassMaster.objects.prefetch_related('sections').filter(status=True)


class ClassSectionMasterListView(generics.ListAPIView):
    """List sections from class_section_master table (fee-backend)"""
    serializer_class = ClassSectionMasterSerializer
    permission_classes = [IsSchoolStaff]

    def get_queryset(self):
        qs = ClassSectionMaster.objects.select_related('class_master').filter(status=True)
        class_id = self.request.query_params.get('class_id')
        if class_id:
            qs = qs.filter(class_master_id=class_id)
        return qs


class SectionListCreateView(generics.ListCreateAPIView):
    serializer_class = SectionSerializer
    permission_classes = [IsSchoolStaff]

    def get_queryset(self):
        qs = Section.objects.all()
        class_id = self.request.query_params.get('class_id')
        if class_id:
            qs = qs.filter(class_ref_id=class_id)
        return qs


class SectionDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = SectionSerializer
    permission_classes = [IsSchoolAdmin]
    queryset = Section.objects.all()


class StudentListCreateView(generics.ListCreateAPIView):
    permission_classes = [IsSchoolStaff]

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return StudentSerializer
        return StudentListSerializer

    def get_queryset(self):
        qs = Student.objects.select_related('class_ref', 'section')
        params = self.request.query_params

        if params.get('class_id'):
            qs = qs.filter(class_ref_id=params['class_id'])
        if params.get('section_id'):
            qs = qs.filter(section_id=params['section_id'])
        if params.get('status'):
            qs = qs.filter(status=params['status'])
        if params.get('search'):
            q = params['search']
            qs = qs.filter(
                Q(first_name__icontains=q) |
                Q(last_name__icontains=q) |
                Q(admission_no__icontains=q) |
                Q(father_name__icontains=q) |
                Q(father_phone__icontains=q)
            )
        return qs


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = StudentSerializer
    permission_classes = [IsSchoolStaff]
    queryset = Student.objects.all()


class StudentStrengthView(APIView):
    permission_classes = [IsSchoolStaff]

    def get(self, request):
        data = (
            Student.objects
            .filter(status='active')
            .values('class_ref__name', 'section__name')
            .annotate(
                total=Count('id'),
                boys=Count('id', filter=Q(gender='M')),
                girls=Count('id', filter=Q(gender='F')),
            )
            .order_by('class_ref__order', 'section__name')
        )
        result = [
            {
                'class_name': row['class_ref__name'],
                'section_name': row['section__name'],
                'total': row['total'],
                'boys': row['boys'],
                'girls': row['girls'],
            }
            for row in data
        ]
        return Response(result)


class ChangeSectionView(APIView):
    permission_classes = [IsSchoolAdmin]

    def post(self, request, pk):
        try:
            student = Student.objects.get(pk=pk)
            section_id = request.data.get('section_id')
            class_id = request.data.get('class_id')
            if section_id:
                student.section_id = section_id
            if class_id:
                student.class_ref_id = class_id
            student.save()
            return Response(StudentSerializer(student).data)
        except Student.DoesNotExist:
            return Response({'detail': 'Student not found.'}, status=status.HTTP_404_NOT_FOUND)


IMPORT_COLUMNS = [
    'admission_no', 'first_name', 'last_name', 'date_of_birth', 'gender',
    'category', 'class_name', 'section_name', 'session_year', 'roll_no',
    'father_name', 'father_phone', 'mother_name', 'mother_phone',
    'address', 'city', 'state', 'pincode', 'admission_date',
    'blood_group', 'aadhar_no',
]


class StudentImportTemplateView(APIView):
    """Download a blank Excel template for student bulk import."""
    permission_classes = [IsSchoolStaff]

    def get(self, request):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Students'

        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)

        for col_idx, col_name in enumerate(IMPORT_COLUMNS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 14)

        # Sample row
        sample = [
            'ADM001', 'Rahul', 'Sharma', '2010-05-15', 'M',
            'GEN', 'Class 1', 'A', '2024-25', '1',
            'Rajesh Sharma', '9876543210', 'Sunita Sharma', '9876543211',
            '123 MG Road', 'Delhi', 'Delhi', '110001', '2024-04-01',
            'A+', '',
        ]
        for col_idx, val in enumerate(sample, 1):
            ws.cell(row=2, column=col_idx, value=val)

        # Notes sheet
        notes = wb.create_sheet('Notes')
        notes['A1'] = 'Field Notes'
        notes['A1'].font = Font(bold=True)
        hints = [
            ('gender', 'M = Male, F = Female, O = Other'),
            ('category', 'GEN / OBC / SC / ST / EWS'),
            ('date_of_birth', 'Format: YYYY-MM-DD  e.g. 2010-05-15'),
            ('admission_date', 'Format: YYYY-MM-DD  e.g. 2024-04-01'),
            ('class_name', 'Must match an existing class e.g. Class 1, Class 2'),
            ('section_name', 'Must match a section under the class e.g. A, B'),
            ('session_year', 'e.g. 2024-25'),
        ]
        for i, (field, hint) in enumerate(hints, 2):
            notes.cell(row=i, column=1, value=field).font = Font(bold=True)
            notes.cell(row=i, column=2, value=hint)

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="student_import_template.xlsx"'
        wb.save(response)
        return response


class StudentBulkImportView(APIView):
    """POST an Excel file to bulk-create students."""
    permission_classes = [IsSchoolStaff]
    parser_classes = [MultiPartParser]

    def post(self, request):
        import openpyxl

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'No file uploaded.'}, status=status.HTTP_400_BAD_REQUEST)
        if not file.name.endswith(('.xlsx', '.xls')):
            return Response({'detail': 'Only .xlsx / .xls files are accepted.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception as e:
            return Response({'detail': f'Could not read file: {e}'}, status=status.HTTP_400_BAD_REQUEST)

        # Read header row
        headers = [str(ws.cell(row=1, column=c).value or '').strip().lower() for c in range(1, ws.max_column + 1)]

        def col(row_values, name):
            try:
                return str(row_values[headers.index(name)] or '').strip()
            except (ValueError, IndexError):
                return ''

        # Cache classes and sections for fast lookup
        class_map = {c.name.lower(): c for c in Class.objects.all()}
        section_map = {}
        for s in Section.objects.select_related('class_ref').all():
            key = (s.class_ref.name.lower(), s.name.lower())
            section_map[key] = s

        created, errors = 0, []

        for row_num in range(2, ws.max_row + 1):
            row_values = [ws.cell(row=row_num, column=c).value for c in range(1, ws.max_column + 1)]
            if all(v is None or str(v).strip() == '' for v in row_values):
                continue  # skip blank rows

            admission_no = col(row_values, 'admission_no')
            first_name   = col(row_values, 'first_name')
            last_name    = col(row_values, 'last_name')
            class_name   = col(row_values, 'class_name')
            section_name = col(row_values, 'section_name')

            if not admission_no or not first_name or not class_name:
                errors.append({'row': row_num, 'error': 'admission_no, first_name, class_name are required'})
                continue

            if Student.objects.filter(admission_no=admission_no).exists():
                errors.append({'row': row_num, 'error': f'admission_no {admission_no} already exists'})
                continue

            class_obj = class_map.get(class_name.lower())
            if not class_obj:
                errors.append({'row': row_num, 'error': f'Class "{class_name}" not found'})
                continue

            section_obj = section_map.get((class_name.lower(), section_name.lower()))

            try:
                Student.objects.create(
                    admission_no=admission_no,
                    first_name=first_name,
                    last_name=last_name or '',
                    date_of_birth=col(row_values, 'date_of_birth') or '2000-01-01',
                    gender=col(row_values, 'gender') or 'M',
                    category=col(row_values, 'category') or 'GEN',
                    class_ref=class_obj,
                    section=section_obj,
                    session_year=col(row_values, 'session_year') or '2024-25',
                    roll_no=col(row_values, 'roll_no'),
                    father_name=col(row_values, 'father_name') or '',
                    father_phone=col(row_values, 'father_phone') or '',
                    mother_name=col(row_values, 'mother_name'),
                    mother_phone=col(row_values, 'mother_phone'),
                    address=col(row_values, 'address') or '',
                    city=col(row_values, 'city'),
                    state=col(row_values, 'state'),
                    pincode=col(row_values, 'pincode'),
                    admission_date=col(row_values, 'admission_date') or '2024-04-01',
                    blood_group=col(row_values, 'blood_group'),
                    aadhar_no=col(row_values, 'aadhar_no'),
                )
                created += 1
            except Exception as e:
                errors.append({'row': row_num, 'error': str(e)})

        return Response({
            'created': created,
            'errors': errors,
            'total_rows': ws.max_row - 1,
        }, status=status.HTTP_200_OK)
